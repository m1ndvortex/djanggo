"""
POS views for ZARGAR jewelry SaaS platform.
Provides touch-optimized POS interface and transaction management.
"""
from django.views.generic import TemplateView, ListView, DetailView, CreateView
from django.views import View
from django.http import JsonResponse, HttpResponse, Http404
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
from django.core.paginator import Paginator
from decimal import Decimal
import json

from django.db import models

from zargar.core.mixins import TenantContextMixin
from .models import POSTransaction, POSTransactionLineItem, POSInvoice, POSOfflineStorage
from .services import POSTransactionService, POSOfflineService, POSInvoiceService, POSReportingService
from zargar.jewelry.models import JewelryItem
from zargar.customers.models import Customer
from zargar.gold_installments.services import GoldPriceService


class POSTransactionListView(LoginRequiredMixin, TenantContextMixin, ListView):
    """List view for POS transactions."""
    
    model = POSTransaction
    template_name = 'pos/transaction_list.html'
    context_object_name = 'transactions'
    paginate_by = 20
    
    def get_queryset(self):
        """Get filtered queryset."""
        queryset = POSTransaction.objects.select_related('customer').order_by('-transaction_date')
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by date range
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(transaction_date__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(transaction_date__date__lte=date_to)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add additional context."""
        context = super().get_context_data(**kwargs)
        context['status_choices'] = POSTransaction.STATUS_CHOICES
        return context


class POSTransactionCreateView(LoginRequiredMixin, TenantContextMixin, TemplateView):
    """Create new POS transaction."""
    
    template_name = 'pos/transaction_create.html'
    
    def post(self, request, *args, **kwargs):
        """Handle transaction creation."""
        try:
            customer_id = request.POST.get('customer_id')
            transaction_type = request.POST.get('transaction_type', 'sale')
            payment_method = request.POST.get('payment_method', 'cash')
            
            transaction = POSTransactionService.create_transaction(
                customer_id=int(customer_id) if customer_id else None,
                transaction_type=transaction_type,
                payment_method=payment_method,
                user=request.user
            )
            
            messages.success(request, _('Transaction created successfully'))
            return redirect('pos:transaction_detail', transaction_id=transaction.transaction_id)
            
        except Exception as e:
            messages.error(request, f'Error creating transaction: {str(e)}')
            return self.get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        """Add context for transaction creation."""
        context = super().get_context_data(**kwargs)
        context['customers'] = Customer.objects.filter(is_active=True)[:100]  # Limit for performance
        context['transaction_types'] = POSTransaction.TRANSACTION_TYPE_CHOICES
        context['payment_methods'] = POSTransaction.PAYMENT_METHOD_CHOICES
        return context


class POSTransactionDetailView(LoginRequiredMixin, TenantContextMixin, DetailView):
    """Detail view for POS transaction."""
    
    model = POSTransaction
    template_name = 'pos/transaction_detail.html'
    context_object_name = 'transaction'
    slug_field = 'transaction_id'
    slug_url_kwarg = 'transaction_id'
    
    def get_context_data(self, **kwargs):
        """Add additional context."""
        context = super().get_context_data(**kwargs)
        transaction = self.get_object()
        
        context['line_items'] = transaction.line_items.select_related('jewelry_item').all()
        context['can_edit'] = transaction.status == 'pending'
        context['current_gold_price'] = GoldPriceService.get_current_gold_price(18)
        
        return context


class POSTransactionCompleteView(LoginRequiredMixin, TenantContextMixin, View):
    """Complete POS transaction."""
    
    def post(self, request, transaction_id):
        """Handle transaction completion."""
        try:
            transaction = get_object_or_404(POSTransaction, transaction_id=transaction_id)
            
            amount_paid = Decimal(request.POST.get('amount_paid', '0'))
            payment_method = request.POST.get('payment_method', transaction.payment_method)
            reference_number = request.POST.get('reference_number', '')
            
            result = POSTransactionService.process_payment(
                transaction=transaction,
                amount_paid=amount_paid,
                payment_method=payment_method,
                reference_number=reference_number
            )
            
            if result['success']:
                messages.success(request, _('Transaction completed successfully'))
                return JsonResponse({
                    'success': True,
                    'change_amount': str(result['change_amount']),
                    'invoice_id': result['invoice'].id
                })
            else:
                return JsonResponse({'success': False, 'error': 'Payment processing failed'})
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class POSTransactionCancelView(LoginRequiredMixin, TenantContextMixin, View):
    """Cancel POS transaction."""
    
    def post(self, request, transaction_id):
        """Handle transaction cancellation."""
        try:
            transaction = get_object_or_404(POSTransaction, transaction_id=transaction_id)
            reason = request.POST.get('reason', '')
            
            result = POSTransactionService.cancel_transaction(
                transaction=transaction,
                reason=reason
            )
            
            if result['success']:
                messages.success(request, _('Transaction cancelled successfully'))
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'error': 'Cancellation failed'})
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class AddJewelryItemView(LoginRequiredMixin, TenantContextMixin, View):
    """Add jewelry item to transaction."""
    
    def post(self, request, transaction_id):
        """Handle adding jewelry item."""
        try:
            transaction = get_object_or_404(POSTransaction, transaction_id=transaction_id)
            
            jewelry_item_id = int(request.POST.get('jewelry_item_id'))
            quantity = int(request.POST.get('quantity', 1))
            custom_price = request.POST.get('custom_price')
            discount_percentage = Decimal(request.POST.get('discount_percentage', '0'))
            
            custom_price = Decimal(custom_price) if custom_price else None
            
            line_item = POSTransactionService.add_jewelry_item_to_transaction(
                transaction=transaction,
                jewelry_item_id=jewelry_item_id,
                quantity=quantity,
                custom_price=custom_price,
                discount_percentage=discount_percentage
            )
            
            return JsonResponse({
                'success': True,
                'line_item_id': line_item.id,
                'subtotal': str(transaction.subtotal),
                'total_amount': str(transaction.total_amount)
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class AddCustomItemView(LoginRequiredMixin, TenantContextMixin, View):
    """Add custom item to transaction."""
    
    def post(self, request, transaction_id):
        """Handle adding custom item."""
        try:
            transaction = get_object_or_404(POSTransaction, transaction_id=transaction_id)
            
            item_name = request.POST.get('item_name')
            unit_price = Decimal(request.POST.get('unit_price'))
            quantity = int(request.POST.get('quantity', 1))
            item_sku = request.POST.get('item_sku', '')
            
            line_item = POSTransactionService.add_custom_item_to_transaction(
                transaction=transaction,
                item_name=item_name,
                unit_price=unit_price,
                quantity=quantity,
                item_sku=item_sku
            )
            
            return JsonResponse({
                'success': True,
                'line_item_id': line_item.id,
                'subtotal': str(transaction.subtotal),
                'total_amount': str(transaction.total_amount)
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class RemoveLineItemView(LoginRequiredMixin, TenantContextMixin, View):
    """Remove line item from transaction."""
    
    def post(self, request, transaction_id, line_item_id):
        """Handle removing line item."""
        try:
            transaction = get_object_or_404(POSTransaction, transaction_id=transaction_id)
            
            POSTransactionService.remove_line_item(
                transaction=transaction,
                line_item_id=line_item_id
            )
            
            return JsonResponse({
                'success': True,
                'subtotal': str(transaction.subtotal),
                'total_amount': str(transaction.total_amount)
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class POSInvoiceDetailView(LoginRequiredMixin, TenantContextMixin, DetailView):
    """Detail view for POS invoice."""
    
    model = POSInvoice
    template_name = 'pos/invoice_detail.html'
    context_object_name = 'invoice'
    pk_url_kwarg = 'invoice_id'


class POSInvoicePDFView(LoginRequiredMixin, TenantContextMixin, View):
    """Generate PDF invoice."""
    
    def get(self, request, invoice_id):
        """Generate and return PDF invoice."""
        try:
            invoice = get_object_or_404(POSInvoice, id=invoice_id)
            
            pdf_content = POSInvoiceService.generate_invoice_pdf(invoice)
            
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="invoice_{invoice.invoice_number}.pdf"'
            
            return response
            
        except Exception as e:
            messages.error(request, f'Error generating PDF: {str(e)}')
            return redirect('pos:invoice_detail', invoice_id=invoice_id)


class POSInvoiceEmailView(LoginRequiredMixin, TenantContextMixin, View):
    """Send invoice via email."""
    
    def post(self, request, invoice_id):
        """Handle sending invoice email."""
        try:
            invoice = get_object_or_404(POSInvoice, id=invoice_id)
            recipient_email = request.POST.get('recipient_email')
            
            if not recipient_email:
                return JsonResponse({'success': False, 'error': 'Email address required'})
            
            success = POSInvoiceService.send_invoice_email(
                invoice=invoice,
                recipient_email=recipient_email,
                include_pdf=True
            )
            
            if success:
                return JsonResponse({'success': True, 'message': 'Invoice sent successfully'})
            else:
                return JsonResponse({'success': False, 'error': 'Failed to send email'})
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class POSOfflineSyncView(LoginRequiredMixin, TenantContextMixin, View):
    """Sync offline POS transactions."""
    
    def post(self, request):
        """Handle offline transaction sync."""
        try:
            device_id = request.POST.get('device_id', '')
            
            sync_results = POSOfflineService.sync_offline_transactions(device_id=device_id)
            
            return JsonResponse({
                'success': True,
                'sync_results': sync_results
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class POSOfflineSyncStatusView(LoginRequiredMixin, TenantContextMixin, View):
    """Get offline sync status."""
    
    def get(self, request):
        """Get sync status summary."""
        try:
            device_id = request.GET.get('device_id', '')
            
            summary = POSOfflineService.get_offline_transaction_summary(device_id=device_id)
            
            return JsonResponse({
                'success': True,
                'summary': summary
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class POSDailySalesReportView(LoginRequiredMixin, TenantContextMixin, TemplateView):
    """Daily sales report view."""
    
    template_name = 'pos/daily_sales_report.html'
    
    def get_context_data(self, **kwargs):
        """Add report data to context."""
        context = super().get_context_data(**kwargs)
        
        date_str = self.request.GET.get('date')
        if date_str:
            from datetime import datetime
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            date = timezone.now().date()
        
        context['report_date'] = date
        context['sales_summary'] = POSReportingService.get_daily_sales_summary(date=date)
        
        return context


class POSMonthlySalesReportView(LoginRequiredMixin, TenantContextMixin, TemplateView):
    """Monthly sales report view."""
    
    template_name = 'pos/monthly_sales_report.html'
    
    def get_context_data(self, **kwargs):
        """Add report data to context."""
        context = super().get_context_data(**kwargs)
        
        year = int(self.request.GET.get('year', timezone.now().year))
        month = int(self.request.GET.get('month', timezone.now().month))
        
        context['report_year'] = year
        context['report_month'] = month
        context['sales_trend'] = POSReportingService.get_monthly_sales_trend(year, month)
        
        return context


# API Views for AJAX/Mobile Integration

class CurrentGoldPriceAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for current gold price."""
    
    def get(self, request):
        """Get current gold price."""
        try:
            karat = int(request.GET.get('karat', 18))
            price_data = GoldPriceService.get_current_gold_price(karat)
            
            return JsonResponse({
                'success': True,
                'price_data': {
                    'price_per_gram': str(price_data['price_per_gram']),
                    'karat': price_data['karat'],
                    'timestamp': price_data['timestamp'].isoformat(),
                    'source': price_data['source']
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class CustomerLookupAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for customer lookup."""
    
    def get(self, request):
        """Search customers."""
        try:
            query = request.GET.get('q', '')
            
            if len(query) < 2:
                return JsonResponse({'success': True, 'customers': []})
            
            customers = Customer.objects.filter(
                models.Q(first_name__icontains=query) |
                models.Q(last_name__icontains=query) |
                models.Q(persian_first_name__icontains=query) |
                models.Q(persian_last_name__icontains=query) |
                models.Q(phone_number__icontains=query)
            )[:10]
            
            customer_data = []
            for customer in customers:
                customer_data.append({
                    'id': customer.id,
                    'name': str(customer),
                    'phone': customer.phone_number,
                    'email': customer.email or '',
                    'loyalty_points': customer.loyalty_points
                })
            
            return JsonResponse({
                'success': True,
                'customers': customer_data
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class JewelryItemSearchAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for jewelry item search."""
    
    def get(self, request):
        """Search jewelry items."""
        try:
            query = request.GET.get('q', '')
            
            if len(query) < 2:
                return JsonResponse({'success': True, 'items': []})
            
            items = JewelryItem.objects.filter(
                models.Q(name__icontains=query) |
                models.Q(sku__icontains=query),
                status='in_stock',
                quantity__gt=0
            ).select_related('category')[:20]
            
            item_data = []
            for item in items:
                item_data.append({
                    'id': item.id,
                    'name': item.name,
                    'sku': item.sku,
                    'category': item.category.name_persian if item.category else '',
                    'weight_grams': str(item.weight_grams),
                    'karat': item.karat,
                    'selling_price': str(item.selling_price or 0),
                    'quantity': item.quantity
                })
            
            return JsonResponse({
                'success': True,
                'items': item_data
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class POSTouchInterfaceView(LoginRequiredMixin, TenantContextMixin, TemplateView):
    """
    Touch-optimized POS interface for tablet use.
    """
    template_name = 'pos/touch_interface.html'
    
    def get_context_data(self, **kwargs):
        """Add context for touch interface."""
        context = super().get_context_data(**kwargs)
        
        # Get current gold price for display
        try:
            current_gold_price = GoldPriceService.get_current_gold_price(18)
            context['current_gold_price'] = current_gold_price
        except Exception:
            context['current_gold_price'] = None
        
        # Get today's stats
        today = timezone.now().date()
        today_transactions = POSTransaction.objects.filter(
            transaction_date__date=today,
            status='completed'
        )
        
        context['today_stats'] = {
            'sales_count': today_transactions.count(),
            'total_sales': today_transactions.aggregate(
                total=models.Sum('total_amount')
            )['total'] or 0,
            'gold_weight': today_transactions.aggregate(
                weight=models.Sum('total_gold_weight_grams')
            )['weight'] or 0,
        }
        
        return context


class POSTodayStatsAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for today's sales statistics."""
    
    def get(self, request):
        """Get today's sales statistics."""
        try:
            today = timezone.now().date()
            today_transactions = POSTransaction.objects.filter(
                transaction_date__date=today,
                status='completed'
            )
            
            stats = {
                'sales_count': today_transactions.count(),
                'total_sales': float(today_transactions.aggregate(
                    total=models.Sum('total_amount')
                )['total'] or 0),
                'gold_weight': float(today_transactions.aggregate(
                    weight=models.Sum('total_gold_weight_grams')
                )['weight'] or 0),
            }
            
            return JsonResponse({
                'success': True,
                'stats': stats
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class POSRecentTransactionsAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for recent transactions."""
    
    def get(self, request):
        """Get recent transactions."""
        try:
            recent_transactions = POSTransaction.objects.select_related('customer').order_by(
                '-transaction_date'
            )[:10]
            
            transactions_data = []
            for transaction in recent_transactions:
                transactions_data.append({
                    'transaction_number': transaction.transaction_number,
                    'customer_name': str(transaction.customer) if transaction.customer else None,
                    'total_amount': float(transaction.total_amount),
                    'transaction_date_shamsi': transaction.transaction_date_shamsi,
                    'status': transaction.status,
                })
            
            return JsonResponse({
                'success': True,
                'transactions': transactions_data
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class POSOfflineTransactionCreateAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for creating offline transactions."""
    
    def post(self, request):
        """Create offline transaction."""
        try:
            from .models import OfflinePOSSystem
            
            data = json.loads(request.body)
            
            device_id = data.get('device_id', '')
            device_name = data.get('device_name', '')
            customer_id = data.get('customer_id')
            line_items = data.get('line_items', [])
            payment_method = data.get('payment_method', 'cash')
            amount_paid = Decimal(str(data.get('amount_paid', '0.00')))
            transaction_type = data.get('transaction_type', 'sale')
            
            # Initialize offline POS system
            offline_pos = OfflinePOSSystem(
                device_id=device_id,
                device_name=device_name
            )
            
            # Create offline transaction
            offline_storage = offline_pos.create_offline_transaction(
                customer_id=customer_id,
                line_items=line_items,
                payment_method=payment_method,
                amount_paid=amount_paid,
                transaction_type=transaction_type
            )
            
            return JsonResponse({
                'success': True,
                'storage_id': str(offline_storage.storage_id),
                'transaction_data': offline_storage.transaction_data,
                'sync_status': offline_storage.sync_status
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)


class POSOfflineSyncAllAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for syncing all offline transactions."""
    
    def post(self, request):
        """Sync all offline transactions for a device."""
        try:
            from .models import OfflinePOSSystem
            
            data = json.loads(request.body)
            device_id = data.get('device_id', '')
            device_name = data.get('device_name', '')
            
            # Initialize offline POS system
            offline_pos = OfflinePOSSystem(
                device_id=device_id,
                device_name=device_name
            )
            
            # Sync offline transactions
            sync_results = offline_pos.sync_offline_transactions()
            
            return JsonResponse({
                'success': True,
                'sync_results': sync_results
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)


class POSOfflineStatusAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for getting offline transaction status."""
    
    def get(self, request):
        """Get offline transaction status for a device."""
        try:
            from .models import OfflinePOSSystem
            
            device_id = request.GET.get('device_id', '')
            device_name = request.GET.get('device_name', '')
            
            # Initialize offline POS system
            offline_pos = OfflinePOSSystem(
                device_id=device_id,
                device_name=device_name
            )
            
            # Get summary
            summary = offline_pos.get_offline_transaction_summary()
            
            return JsonResponse({
                'success': True,
                'summary': summary
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)


class POSOfflineConflictResolveAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for resolving offline sync conflicts."""
    
    def post(self, request):
        """Resolve offline sync conflicts."""
        try:
            from .models import OfflinePOSSystem
            
            data = json.loads(request.body)
            device_id = data.get('device_id', '')
            device_name = data.get('device_name', '')
            resolution_actions = data.get('resolution_actions', {})
            
            # Initialize offline POS system
            offline_pos = OfflinePOSSystem(
                device_id=device_id,
                device_name=device_name
            )
            
            # Resolve conflicts
            results = offline_pos.resolve_sync_conflicts(resolution_actions)
            
            return JsonResponse({
                'success': True,
                'resolution_results': results
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)


class POSOfflineCleanupAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for cleaning up old offline transactions."""
    
    def post(self, request):
        """Clean up old synced offline transactions."""
        try:
            from .models import OfflinePOSSystem
            
            data = json.loads(request.body)
            device_id = data.get('device_id', '')
            device_name = data.get('device_name', '')
            days_old = int(data.get('days_old', 30))
            
            # Initialize offline POS system
            offline_pos = OfflinePOSSystem(
                device_id=device_id,
                device_name=device_name
            )
            
            # Clean up old transactions
            cleaned_count = offline_pos.cleanup_old_transactions(days_old=days_old)
            
            return JsonResponse({
                'success': True,
                'cleaned_count': cleaned_count
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)


class POSOfflineExportAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for exporting offline transaction data."""
    
    def get(self, request):
        """Export offline transaction data for backup."""
        try:
            from .models import OfflinePOSSystem
            
            device_id = request.GET.get('device_id', '')
            device_name = request.GET.get('device_name', '')
            
            # Initialize offline POS system
            offline_pos = OfflinePOSSystem(
                device_id=device_id,
                device_name=device_name
            )
            
            # Export data
            export_data = offline_pos.export_offline_data()
            
            return JsonResponse({
                'success': True,
                'export_data': export_data
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)


class POSCreateCustomerAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for creating new customers."""
    
    def post(self, request):
        """Create new customer."""
        try:
            import json
            from zargar.customers.models import Customer
            
            data = json.loads(request.body)
            
            # Validate required fields
            if not data.get('persian_first_name') or not data.get('persian_last_name'):
                return JsonResponse({
                    'success': False,
                    'errors': ['نام و نام خانوادگی فارسی الزامی است']
                })
            
            if not data.get('phone_number'):
                return JsonResponse({
                    'success': False,
                    'errors': ['شماره تلفن الزامی است']
                })
            
            # Check if phone number already exists
            if Customer.objects.filter(phone_number=data['phone_number']).exists():
                return JsonResponse({
                    'success': False,
                    'errors': ['مشتری با این شماره تلفن قبلاً ثبت شده است']
                })
            
            # Create customer
            customer = Customer.objects.create(
                persian_first_name=data['persian_first_name'],
                persian_last_name=data['persian_last_name'],
                first_name=data.get('first_name', ''),
                last_name=data.get('last_name', ''),
                phone_number=data['phone_number'],
                email=data.get('email', ''),
                birth_date_shamsi=data.get('birth_date_shamsi', ''),
                national_id=data.get('national_id', ''),
                address=data.get('address', ''),
                customer_type=data.get('customer_type', 'individual'),
                created_by=request.user
            )
            
            return JsonResponse({
                'success': True,
                'customer': {
                    'id': customer.id,
                    'full_name': customer.full_name,
                    'full_persian_name': customer.full_persian_name,
                    'phone_number': customer.phone_number,
                    'email': customer.email,
                    'is_vip': customer.is_vip,
                    'loyalty_points': customer.loyalty_points,
                    'total_purchases': float(customer.total_purchases),
                    'last_purchase_date_shamsi': customer.last_purchase_date.strftime('%Y/%m/%d') if customer.last_purchase_date else None
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'errors': [str(e)]
            }, status=400)


class POSPaymentHistoryAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for customer payment history."""
    
    def get(self, request):
        """Get customer payment history."""
        try:
            from zargar.customers.models import Customer
            from django.core.paginator import Paginator
            from datetime import datetime, timedelta
            
            customer_id = request.GET.get('customer_id')
            if not customer_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Customer ID required'
                })
            
            customer = get_object_or_404(Customer, id=customer_id)
            
            # Get filter parameters
            period = request.GET.get('period', 'all')
            payment_type = request.GET.get('type', 'all')
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 10))
            
            # Build queryset
            transactions = POSTransaction.objects.filter(
                customer=customer,
                status='completed'
            ).select_related('invoice').order_by('-transaction_date')
            
            # Apply period filter
            if period != 'all':
                today = timezone.now().date()
                if period == 'month':
                    start_date = today.replace(day=1)
                elif period == 'quarter':
                    start_date = today - timedelta(days=90)
                elif period == 'year':
                    start_date = today.replace(month=1, day=1)
                
                transactions = transactions.filter(transaction_date__date__gte=start_date)
            
            # Apply payment type filter
            if payment_type != 'all':
                transactions = transactions.filter(payment_method=payment_type)
            
            # Calculate summary
            summary = {
                'total_payments': transactions.aggregate(
                    total=models.Sum('total_amount')
                )['total'] or 0,
                'cash_payments': transactions.filter(payment_method='cash').aggregate(
                    total=models.Sum('total_amount')
                )['total'] or 0,
                'card_payments': transactions.filter(payment_method='card').aggregate(
                    total=models.Sum('total_amount')
                )['total'] or 0,
                'remaining_debt': customer.current_debt if hasattr(customer, 'current_debt') else 0
            }
            
            # Paginate
            paginator = Paginator(transactions, page_size)
            page_obj = paginator.get_page(page)
            
            # Format payment data
            payments = []
            for transaction in page_obj:
                from zargar.core.calendar_utils import PersianCalendarUtils
                
                shamsi_date = PersianCalendarUtils.gregorian_to_shamsi(transaction.transaction_date.date())
                date_shamsi = f"{shamsi_date[0]:04d}/{shamsi_date[1]:02d}/{shamsi_date[2]:02d}"
                
                payments.append({
                    'id': transaction.id,
                    'date_shamsi': date_shamsi,
                    'invoice_number': transaction.invoice.invoice_number if hasattr(transaction, 'invoice') else transaction.transaction_number,
                    'invoice_url': f"/pos/invoice/{transaction.invoice.id}/" if hasattr(transaction, 'invoice') else '#',
                    'invoice_pdf_url': f"/pos/invoice/{transaction.invoice.id}/pdf/" if hasattr(transaction, 'invoice') else '#',
                    'payment_method': transaction.payment_method,
                    'payment_method_display': transaction.get_payment_method_display(),
                    'amount': float(transaction.total_amount),
                    'status': transaction.status,
                    'status_display': transaction.get_status_display()
                })
            
            return JsonResponse({
                'success': True,
                'payments': payments,
                'summary': {
                    'total_payments': float(summary['total_payments']),
                    'cash_payments': float(summary['cash_payments']),
                    'card_payments': float(summary['card_payments']),
                    'remaining_debt': float(summary['remaining_debt'])
                },
                'total_pages': paginator.num_pages,
                'total_records': paginator.count,
                'current_page': page
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)


class POSPaymentHistoryExportAPIView(LoginRequiredMixin, TenantContextMixin, View):
    """API endpoint for exporting payment history."""
    
    def get(self, request):
        """Export payment history to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from django.http import HttpResponse
            from zargar.customers.models import Customer
            from datetime import datetime, timedelta
            
            customer_id = request.GET.get('customer_id')
            if not customer_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Customer ID required'
                })
            
            customer = get_object_or_404(Customer, id=customer_id)
            
            # Get filter parameters
            period = request.GET.get('period', 'all')
            payment_type = request.GET.get('type', 'all')
            
            # Build queryset
            transactions = POSTransaction.objects.filter(
                customer=customer,
                status='completed'
            ).select_related('invoice').order_by('-transaction_date')
            
            # Apply filters (same logic as above)
            if period != 'all':
                today = timezone.now().date()
                if period == 'month':
                    start_date = today.replace(day=1)
                elif period == 'quarter':
                    start_date = today - timedelta(days=90)
                elif period == 'year':
                    start_date = today.replace(month=1, day=1)
                
                transactions = transactions.filter(transaction_date__date__gte=start_date)
            
            if payment_type != 'all':
                transactions = transactions.filter(payment_method=payment_type)
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Payment History"
            
            # Headers
            headers = ['تاریخ', 'شماره فاکتور', 'نوع پرداخت', 'مبلغ', 'وضعیت']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for row, transaction in enumerate(transactions, 2):
                from zargar.core.calendar_utils import PersianCalendarUtils
                
                shamsi_date = PersianCalendarUtils.gregorian_to_shamsi(transaction.transaction_date.date())
                date_shamsi = f"{shamsi_date[0]:04d}/{shamsi_date[1]:02d}/{shamsi_date[2]:02d}"
                
                ws.cell(row=row, column=1, value=date_shamsi)
                ws.cell(row=row, column=2, value=transaction.invoice.invoice_number if hasattr(transaction, 'invoice') else transaction.transaction_number)
                ws.cell(row=row, column=3, value=transaction.get_payment_method_display())
                ws.cell(row=row, column=4, value=float(transaction.total_amount))
                ws.cell(row=row, column=5, value=transaction.get_status_display())
            
            # Prepare response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="payment_history_{customer.phone_number}.xlsx"'
            
            wb.save(response)
            return response
            
        except ImportError:
            return JsonResponse({
                'success': False,
                'error': 'Excel export not available'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
            
            return JsonResponse({
                'success': True,
                'export_data': export_data
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)


class PlaceholderView(TemplateView):
    """
    Placeholder view for POS module (backward compatibility).
    """
    def get(self, request, *args, **kwargs):
        return JsonResponse({
            'module': 'pos',
            'status': 'placeholder',
            'message': 'POS module will be implemented in later tasks'
        })