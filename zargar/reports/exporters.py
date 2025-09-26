"""
Report export functionality for ZARGAR jewelry SaaS platform.

This module handles exporting reports to various formats including PDF, Excel, CSV, and JSON
with proper Persian formatting and RTL layout support.
"""

import os
import json
import csv
from io import BytesIO, StringIO
from typing import Dict, Any, Optional
from decimal import Decimal
from datetime import date, datetime

from django.conf import settings
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.utils import timezone

# PDF generation
try:
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

# Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportExporter:
    """
    Main class for exporting reports to various formats.
    """
    
    def __init__(self):
        """Initialize the report exporter."""
        self.reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def export_report(self, report_data: Dict[str, Any], output_format: str, filename: str) -> str:
        """
        Export report data to specified format.
        
        Args:
            report_data: Report data dictionary
            output_format: Output format (pdf, excel, csv, json)
            filename: Output filename
            
        Returns:
            Path to generated file
        """
        if output_format == 'pdf':
            return self.export_to_pdf(report_data, filename)
        elif output_format == 'excel':
            return self.export_to_excel(report_data, filename)
        elif output_format == 'csv':
            return self.export_to_csv(report_data, filename)
        elif output_format == 'json':
            return self.export_to_json(report_data, filename)
        else:
            raise ValueError(f"Unsupported output format: {output_format}")
    
    def export_to_pdf(self, report_data: Dict[str, Any], filename: str) -> str:
        """
        Export report to PDF format with Persian RTL layout.
        
        Args:
            report_data: Report data dictionary
            filename: Output filename
            
        Returns:
            Path to generated PDF file
        """
        if not WEASYPRINT_AVAILABLE:
            raise ImportError("WeasyPrint is required for PDF export")
        
        # Determine template based on report type
        report_type = report_data.get('report_type', 'generic')
        template_name = f'reports/pdf/{report_type}.html'
        
        # Fallback to generic template if specific template doesn't exist
        try:
            html_content = render_to_string(template_name, {'report': report_data})
        except:
            html_content = render_to_string('reports/pdf/generic.html', {'report': report_data})
        
        # Generate PDF
        font_config = FontConfiguration()
        
        # CSS for Persian RTL layout
        css_content = """
        @font-face {
            font-family: 'Vazir';
            src: url('static/fonts/Vazir-Regular.woff2') format('woff2');
        }
        
        body {
            font-family: 'Vazir', Arial, sans-serif;
            direction: rtl;
            text-align: right;
            font-size: 12px;
            line-height: 1.6;
        }
        
        .report-header {
            text-align: center;
            margin-bottom: 30px;
            border-bottom: 2px solid #333;
            padding-bottom: 15px;
        }
        
        .report-title {
            font-size: 18px;
            font-weight: bold;
            margin-bottom: 10px;
        }
        
        .report-date {
            font-size: 12px;
            color: #666;
        }
        
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            direction: rtl;
        }
        
        th, td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: right;
        }
        
        th {
            background-color: #f5f5f5;
            font-weight: bold;
        }
        
        .number {
            text-align: left;
            font-family: 'Vazir', monospace;
        }
        
        .total-row {
            font-weight: bold;
            background-color: #f9f9f9;
        }
        
        .report-footer {
            margin-top: 30px;
            text-align: center;
            font-size: 10px;
            color: #666;
        }
        """
        
        css = CSS(string=css_content, font_config=font_config)
        
        # Generate PDF file
        file_path = os.path.join(self.reports_dir, filename)
        HTML(string=html_content).write_pdf(file_path, stylesheets=[css], font_config=font_config)
        
        return file_path
    
    def export_to_excel(self, report_data: Dict[str, Any], filename: str) -> str:
        """
        Export report to Excel format with Persian formatting.
        
        Args:
            report_data: Report data dictionary
            filename: Output filename
            
        Returns:
            Path to generated Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Set worksheet title
        report_title = report_data.get('report_title_persian', 'گزارش')
        ws.title = report_title[:31]  # Excel sheet name limit
        
        # Configure RTL
        ws.sheet_view.rightToLeft = True
        
        # Define styles
        header_font = Font(name='Tahoma', size=14, bold=True)
        title_font = Font(name='Tahoma', size=16, bold=True)
        data_font = Font(name='Tahoma', size=11)
        number_font = Font(name='Tahoma', size=11)
        
        header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add report header
        current_row = 1
        
        # Report title
        ws.merge_cells(f'A{current_row}:F{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = report_title
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 2
        
        # Report date range
        if report_data.get('date_from_shamsi') and report_data.get('date_to_shamsi'):
            ws.merge_cells(f'A{current_row}:F{current_row}')
            date_cell = ws[f'A{current_row}']
            date_cell.value = f"از تاریخ {report_data['date_from_shamsi']} تا {report_data['date_to_shamsi']}"
            date_cell.font = data_font
            date_cell.alignment = Alignment(horizontal='center')
            current_row += 2
        
        # Export based on report type
        report_type = report_data.get('report_type')
        
        if report_type == 'trial_balance':
            current_row = self._export_trial_balance_to_excel(ws, report_data, current_row)
        elif report_type == 'profit_loss':
            current_row = self._export_profit_loss_to_excel(ws, report_data, current_row)
        elif report_type == 'balance_sheet':
            current_row = self._export_balance_sheet_to_excel(ws, report_data, current_row)
        elif report_type == 'inventory_valuation':
            current_row = self._export_inventory_valuation_to_excel(ws, report_data, current_row)
        elif report_type == 'customer_aging':
            current_row = self._export_customer_aging_to_excel(ws, report_data, current_row)
        else:
            # Generic export
            current_row = self._export_generic_to_excel(ws, report_data, current_row)
        
        # Add generation timestamp
        current_row += 2
        ws[f'A{current_row}'] = f"تاریخ تولید گزارش: {report_data.get('generated_at_shamsi', '')}"
        ws[f'A{current_row}'].font = Font(name='Tahoma', size=9, italic=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        file_path = os.path.join(self.reports_dir, filename)
        wb.save(file_path)
        
        return file_path
    
    def _export_trial_balance_to_excel(self, ws, report_data: Dict[str, Any], start_row: int) -> int:
        """Export trial balance data to Excel worksheet."""
        current_row = start_row
        
        # Headers
        headers = ['کد حساب', 'نام حساب', 'نوع حساب', 'بدهکار', 'بستانکار']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(name='Tahoma', size=12, bold=True)
            cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # Data rows
        for account in report_data.get('accounts', []):
            ws.cell(row=current_row, column=1, value=account['account_code'])
            ws.cell(row=current_row, column=2, value=account['account_name_persian'])
            ws.cell(row=current_row, column=3, value=account['account_type'])
            ws.cell(row=current_row, column=4, value=account['debit_amount_formatted'])
            ws.cell(row=current_row, column=5, value=account['credit_amount_formatted'])
            current_row += 1
        
        # Totals row
        current_row += 1
        ws.cell(row=current_row, column=1, value='جمع کل').font = Font(bold=True)
        ws.cell(row=current_row, column=4, value=report_data.get('total_debits_formatted', '')).font = Font(bold=True)
        ws.cell(row=current_row, column=5, value=report_data.get('total_credits_formatted', '')).font = Font(bold=True)
        
        return current_row + 2
    
    def _export_profit_loss_to_excel(self, ws, report_data: Dict[str, Any], start_row: int) -> int:
        """Export profit & loss data to Excel worksheet."""
        current_row = start_row
        
        # Revenue section
        ws.cell(row=current_row, column=1, value='درآمدها').font = Font(name='Tahoma', size=14, bold=True)
        current_row += 1
        
        for account in report_data.get('revenue_accounts', []):
            ws.cell(row=current_row, column=1, value=account['account_name_persian'])
            ws.cell(row=current_row, column=2, value=account['amount_formatted'])
            current_row += 1
        
        ws.cell(row=current_row, column=1, value='جمع درآمدها').font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=report_data.get('total_revenue_formatted', '')).font = Font(bold=True)
        current_row += 2
        
        # Expense section
        ws.cell(row=current_row, column=1, value='هزینه‌ها').font = Font(name='Tahoma', size=14, bold=True)
        current_row += 1
        
        for account in report_data.get('expense_accounts', []):
            ws.cell(row=current_row, column=1, value=account['account_name_persian'])
            ws.cell(row=current_row, column=2, value=account['amount_formatted'])
            current_row += 1
        
        ws.cell(row=current_row, column=1, value='جمع هزینه‌ها').font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=report_data.get('total_expenses_formatted', '')).font = Font(bold=True)
        current_row += 2
        
        # Net income
        ws.cell(row=current_row, column=1, value='سود خالص').font = Font(name='Tahoma', size=14, bold=True)
        ws.cell(row=current_row, column=2, value=report_data.get('net_income_formatted', '')).font = Font(name='Tahoma', size=14, bold=True)
        
        return current_row + 2
    
    def _export_balance_sheet_to_excel(self, ws, report_data: Dict[str, Any], start_row: int) -> int:
        """Export balance sheet data to Excel worksheet."""
        current_row = start_row
        
        # Assets section
        ws.cell(row=current_row, column=1, value='دارایی‌ها').font = Font(name='Tahoma', size=14, bold=True)
        current_row += 1
        
        for category, accounts in report_data.get('assets', {}).items():
            ws.cell(row=current_row, column=1, value=category).font = Font(bold=True)
            current_row += 1
            
            for account in accounts:
                ws.cell(row=current_row, column=2, value=account['account_name_persian'])
                ws.cell(row=current_row, column=3, value=account['amount_formatted'])
                current_row += 1
        
        ws.cell(row=current_row, column=1, value='جمع دارایی‌ها').font = Font(bold=True)
        ws.cell(row=current_row, column=3, value=report_data.get('total_assets_formatted', '')).font = Font(bold=True)
        current_row += 2
        
        # Liabilities section
        ws.cell(row=current_row, column=1, value='بدهی‌ها').font = Font(name='Tahoma', size=14, bold=True)
        current_row += 1
        
        for category, accounts in report_data.get('liabilities', {}).items():
            ws.cell(row=current_row, column=1, value=category).font = Font(bold=True)
            current_row += 1
            
            for account in accounts:
                ws.cell(row=current_row, column=2, value=account['account_name_persian'])
                ws.cell(row=current_row, column=3, value=account['amount_formatted'])
                current_row += 1
        
        # Equity section
        ws.cell(row=current_row, column=1, value='حقوق صاحبان سهام').font = Font(name='Tahoma', size=14, bold=True)
        current_row += 1
        
        for category, accounts in report_data.get('equity', {}).items():
            ws.cell(row=current_row, column=1, value=category).font = Font(bold=True)
            current_row += 1
            
            for account in accounts:
                ws.cell(row=current_row, column=2, value=account['account_name_persian'])
                ws.cell(row=current_row, column=3, value=account['amount_formatted'])
                current_row += 1
        
        ws.cell(row=current_row, column=1, value='جمع بدهی‌ها و حقوق صاحبان سهام').font = Font(bold=True)
        ws.cell(row=current_row, column=3, value=report_data.get('total_liabilities_equity_formatted', '')).font = Font(bold=True)
        
        return current_row + 2
    
    def _export_inventory_valuation_to_excel(self, ws, report_data: Dict[str, Any], start_row: int) -> int:
        """Export inventory valuation data to Excel worksheet."""
        current_row = start_row
        
        # Headers
        headers = ['کد کالا', 'نام کالا', 'تعداد', 'وزن (گرم)', 'عیار', 'اجرت ساخت', 'ارزش کل']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(name='Tahoma', size=12, bold=True)
            cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        current_row += 1
        
        # Data by category
        for category in report_data.get('categories', []):
            # Category header
            ws.cell(row=current_row, column=1, value=category['category_name']).font = Font(bold=True)
            current_row += 1
            
            # Category items
            for item in category['items']:
                ws.cell(row=current_row, column=1, value=item['sku'])
                ws.cell(row=current_row, column=2, value=item['name'])
                ws.cell(row=current_row, column=3, value=item['quantity'])
                ws.cell(row=current_row, column=4, value=item['weight_grams_formatted'])
                ws.cell(row=current_row, column=5, value=item['karat'])
                ws.cell(row=current_row, column=6, value=item['manufacturing_cost_formatted'])
                ws.cell(row=current_row, column=7, value=item['total_item_value_formatted'])
                current_row += 1
            
            # Category total
            ws.cell(row=current_row, column=6, value='جمع دسته‌بندی:').font = Font(bold=True)
            ws.cell(row=current_row, column=7, value=category['category_total_value_formatted']).font = Font(bold=True)
            current_row += 2
        
        # Grand total
        ws.cell(row=current_row, column=6, value='جمع کل موجودی:').font = Font(name='Tahoma', size=14, bold=True)
        ws.cell(row=current_row, column=7, value=report_data.get('total_inventory_value_formatted', '')).font = Font(name='Tahoma', size=14, bold=True)
        
        return current_row + 2
    
    def _export_customer_aging_to_excel(self, ws, report_data: Dict[str, Any], start_row: int) -> int:
        """Export customer aging data to Excel worksheet."""
        current_row = start_row
        
        # Headers
        headers = ['نام مشتری', 'شماره تماس', 'جمع مطالبات'] + report_data.get('aging_period_labels', [])
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(name='Tahoma', size=12, bold=True)
            cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        current_row += 1
        
        # Customer data
        for customer in report_data.get('customers', []):
            ws.cell(row=current_row, column=1, value=customer['customer_name'])
            ws.cell(row=current_row, column=2, value=customer['phone_number'])
            ws.cell(row=current_row, column=3, value=customer['total_outstanding_formatted'])
            
            # Aging breakdown
            col = 4
            for period_key in customer['aging_breakdown_formatted']:
                ws.cell(row=current_row, column=col, value=customer['aging_breakdown_formatted'][period_key])
                col += 1
            
            current_row += 1
        
        # Totals row
        current_row += 1
        ws.cell(row=current_row, column=1, value='جمع کل').font = Font(bold=True)
        ws.cell(row=current_row, column=3, value=report_data.get('total_outstanding_formatted', '')).font = Font(bold=True)
        
        col = 4
        for period_key in report_data.get('aging_totals_formatted', {}):
            ws.cell(row=current_row, column=col, value=report_data['aging_totals_formatted'][period_key]).font = Font(bold=True)
            col += 1
        
        return current_row + 2
    
    def _export_generic_to_excel(self, ws, report_data: Dict[str, Any], start_row: int) -> int:
        """Export generic report data to Excel worksheet."""
        current_row = start_row
        
        # Add key-value pairs from report data
        for key, value in report_data.items():
            if key not in ['report_type', 'generated_at', 'generated_at_shamsi']:
                ws.cell(row=current_row, column=1, value=str(key))
                ws.cell(row=current_row, column=2, value=str(value))
                current_row += 1
        
        return current_row + 2
    
    def export_to_csv(self, report_data: Dict[str, Any], filename: str) -> str:
        """
        Export report to CSV format.
        
        Args:
            report_data: Report data dictionary
            filename: Output filename
            
        Returns:
            Path to generated CSV file
        """
        file_path = os.path.join(self.reports_dir, filename)
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            writer.writerow([
                report_data.get('report_title_persian', 'گزارش'),
                f"تاریخ: {report_data.get('generated_at_shamsi', '')}"
            ])
            writer.writerow([])  # Empty row
            
            # Export based on report type
            report_type = report_data.get('report_type')
            
            if report_type == 'trial_balance':
                self._export_trial_balance_to_csv(writer, report_data)
            elif report_type == 'inventory_valuation':
                self._export_inventory_valuation_to_csv(writer, report_data)
            elif report_type == 'customer_aging':
                self._export_customer_aging_to_csv(writer, report_data)
            else:
                # Generic CSV export
                for key, value in report_data.items():
                    if isinstance(value, (str, int, float, Decimal)):
                        writer.writerow([key, value])
        
        return file_path
    
    def _export_trial_balance_to_csv(self, writer, report_data: Dict[str, Any]):
        """Export trial balance to CSV."""
        writer.writerow(['کد حساب', 'نام حساب', 'نوع حساب', 'بدهکار', 'بستانکار'])
        
        for account in report_data.get('accounts', []):
            writer.writerow([
                account['account_code'],
                account['account_name_persian'],
                account['account_type'],
                account['debit_amount_formatted'],
                account['credit_amount_formatted']
            ])
        
        writer.writerow([])
        writer.writerow(['جمع کل', '', '', 
                        report_data.get('total_debits_formatted', ''),
                        report_data.get('total_credits_formatted', '')])
    
    def _export_inventory_valuation_to_csv(self, writer, report_data: Dict[str, Any]):
        """Export inventory valuation to CSV."""
        writer.writerow(['کد کالا', 'نام کالا', 'تعداد', 'وزن (گرم)', 'عیار', 'ارزش کل'])
        
        for category in report_data.get('categories', []):
            writer.writerow([f"دسته‌بندی: {category['category_name']}"])
            
            for item in category['items']:
                writer.writerow([
                    item['sku'],
                    item['name'],
                    item['quantity'],
                    item['weight_grams_formatted'],
                    item['karat'],
                    item['total_item_value_formatted']
                ])
            
            writer.writerow(['', '', '', '', 'جمع دسته‌بندی:', category['category_total_value_formatted']])
            writer.writerow([])
        
        writer.writerow(['', '', '', '', 'جمع کل:', report_data.get('total_inventory_value_formatted', '')])
    
    def _export_customer_aging_to_csv(self, writer, report_data: Dict[str, Any]):
        """Export customer aging to CSV."""
        headers = ['نام مشتری', 'شماره تماس', 'جمع مطالبات'] + report_data.get('aging_period_labels', [])
        writer.writerow(headers)
        
        for customer in report_data.get('customers', []):
            row = [
                customer['customer_name'],
                customer['phone_number'],
                customer['total_outstanding_formatted']
            ]
            
            for period_key in customer['aging_breakdown_formatted']:
                row.append(customer['aging_breakdown_formatted'][period_key])
            
            writer.writerow(row)
        
        # Totals row
        totals_row = ['جمع کل', '', report_data.get('total_outstanding_formatted', '')]
        for period_key in report_data.get('aging_totals_formatted', {}):
            totals_row.append(report_data['aging_totals_formatted'][period_key])
        
        writer.writerow(totals_row)
    
    def export_to_json(self, report_data: Dict[str, Any], filename: str) -> str:
        """
        Export report to JSON format.
        
        Args:
            report_data: Report data dictionary
            filename: Output filename
            
        Returns:
            Path to generated JSON file
        """
        file_path = os.path.join(self.reports_dir, filename)
        
        # Convert datetime and date objects to strings for JSON serialization
        json_data = self._prepare_for_json(report_data)
        
        with open(file_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(json_data, jsonfile, ensure_ascii=False, indent=2, default=str)
        
        return file_path
    
    def _prepare_for_json(self, data):
        """Prepare data for JSON serialization by converting non-serializable types."""
        if isinstance(data, dict):
            return {key: self._prepare_for_json(value) for key, value in data.items()}
        elif isinstance(data, list):
            return [self._prepare_for_json(item) for item in data]
        elif isinstance(data, (datetime, date)):
            return data.isoformat()
        elif isinstance(data, Decimal):
            return float(data)
        else:
            return data
    
    def create_http_response(self, report_data: Dict[str, Any], output_format: str, filename: str) -> HttpResponse:
        """
        Create HTTP response for report download.
        
        Args:
            report_data: Report data dictionary
            output_format: Output format
            filename: Download filename
            
        Returns:
            HttpResponse for file download
        """
        if output_format == 'json':
            # Return JSON directly in response
            json_data = self._prepare_for_json(report_data)
            response = HttpResponse(
                json.dumps(json_data, ensure_ascii=False, indent=2, default=str),
                content_type='application/json; charset=utf-8'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
        else:
            # Generate file and return as download
            file_path = self.export_report(report_data, output_format, filename)
            
            with open(file_path, 'rb') as f:
                file_content = f.read()
            
            # Determine content type
            content_types = {
                'pdf': 'application/pdf',
                'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'csv': 'text/csv; charset=utf-8-sig',
            }
            
            response = HttpResponse(
                file_content,
                content_type=content_types.get(output_format, 'application/octet-stream')
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response